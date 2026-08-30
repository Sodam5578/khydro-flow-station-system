import os
import json
import pandas as pd
import openpyxl

def dms_to_dd(dms_str):
    if pd.isna(dms_str) or not str(dms_str).strip():
        return None
    try:
        parts = [float(p.strip()) for p in str(dms_str).strip().split("-") if p.strip()]
        if len(parts) == 3:
            return round(parts[0] + parts[1]/60.0 + parts[2]/3600.0, 6)
        elif len(parts) == 2:
            return round(parts[0] + parts[1]/60.0, 6)
        elif len(parts) == 1:
            return round(parts[0], 6)
    except Exception as e:
        return None
    return None

def clean_val(val):
    if pd.isna(val) or val is None or str(val).strip() == "nan":
        return ""
    val_str = str(val).strip()
    if val_str.endswith(".0") and val_str[:-2].isdigit():
        return val_str[:-2]
    return val_str

def get_gauge_category(gauge_type, advm_cnt, ewsv_cnt):
    gt = str(gauge_type).upper()
    has_ewsv = "EWSV" in gt or (ewsv_cnt not in ["", "-", "0", "nan"])
    has_advm = "ADVM" in gt or (advm_cnt not in ["", "-", "0", "nan"])
    
    if ("EWSV" in gt and "ADVM" in gt) or (has_ewsv and has_advm and "EWSV" in gt and "ADVM" in gt):
        return "DUAL"
    elif "EWSV" in gt:
        return "EWSV"
    elif "ADVM" in gt:
        return "ADVM"
    return "OTHER"

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    system_dir = os.path.dirname(script_dir)
    workspace_dir = os.path.dirname(system_dir)

    excel_main = os.path.join(workspace_dir, "자동유량측정시설현황.xlsx")
    excel_maint = os.path.join(workspace_dir, "2026년 자동유량측정시설 유지관리 지점별 필요사항 현황.xlsx")

    print(f"Loading main excel: {excel_main}")
    df_main = pd.read_excel(excel_main)

    maint_dict = {}
    if os.path.exists(excel_maint):
        try:
            wb = openpyxl.load_workbook(excel_maint, data_only=True)
            if "지점 유지관리" in wb.sheetnames:
                ws = wb["지점 유지관리"]
                headers = [ws.cell(2, c).value for c in range(1, ws.max_column+1)]
                for r in range(3, ws.max_row+1):
                    code_val = ws.cell(r, 5).value
                    name_val = ws.cell(r, 6).value
                    if code_val or name_val:
                        key = str(code_val).strip() if code_val else str(name_val).strip()
                        maint_info = {}
                        for c in range(1, ws.max_column+1):
                            h = headers[c-1]
                            v = ws.cell(r, c).value
                            if h and v is not None:
                                h_clean = str(h).replace("\n", " ").strip()
                                maint_info[h_clean] = str(v).strip()
                        maint_dict[key] = maint_info
                print(f"Loaded maintenance info for {len(maint_dict)} stations.")
        except Exception as e:
            print(f"Warning: Could not parse maintenance excel: {e}")

    stations = []
    for idx, row in df_main.iterrows():
        lon_raw = clean_val(row.get("경도"))
        lat_raw = clean_val(row.get("위도"))
        lon_dd = dms_to_dd(lon_raw)
        lat_dd = dms_to_dd(lat_raw)
        
        station_code = clean_val(row.get("지점코드"))
        station_name = clean_val(row.get("지점명"))
        gauge_type_raw = clean_val(row.get("형식(유속계)"))
        advm_count = clean_val(row.get("ADVM (대)"))
        ewsv_count = clean_val(row.get("EWSV (대)"))

        gauge_category = get_gauge_category(gauge_type_raw, advm_count, ewsv_count)
        is_dual_gauge = (gauge_category == "DUAL")

        if station_name == "충주시(국원대교)" and lat_raw == "35-57-58":
            lat_raw = "36-57-58"
            lat_dd = dms_to_dd(lat_raw)
        m_info = maint_dict.get(station_code, maint_dict.get(station_name, {}))

        station_obj = {
            "id": idx + 1,
            "seq": int(row.get("연번")) if pd.notna(row.get("연번")) else (idx + 1),
            "region": clean_val(row.get("관할명")),
            "river": clean_val(row.get("하천명")),
            "name": station_name,
            "address": clean_val(row.get("위치")),
            "code": station_code,
            "installYear": clean_val(row.get("설치년도")),
            "obsStartYear": clean_val(row.get("관측개시년도")),
            "isOperating2026": clean_val(row.get("2026년 운영")) in ["○", "O", "o", "1", "Y", "y", "운영"],
            "installDirection": clean_val(row.get("설치방향")),
            "floodAlert": clean_val(row.get("홍수특보")) in ["○", "O", "o", "1", "Y", "y"],
            "droughtAlert": clean_val(row.get("갈수예보")) in ["○", "O", "o", "1", "Y", "y"],
            "drainage": clean_val(row.get("배수")) in ["○", "O", "o", "1", "Y", "y"],
            "tide": clean_val(row.get("조위")) in ["○", "O", "o", "1", "Y", "y"],
            "pollutionTotal": clean_val(row.get("오염총량")) in ["○", "O", "o", "1", "Y", "y"],
            "gaugeType": gauge_type_raw,
            "gaugeCategory": gauge_category,
            "isDualGauge": is_dual_gauge,
            "advmCount": advm_count,
            "ewsvCount": ewsv_count,
            "calib2026": clean_val(row.get("26년 유속계 검정지점")) in ["○", "O", "o", "1", "Y", "y"],
            "calibCount2026": clean_val(row.get("26년 검정대상 유속계(대)")),
            "solarInstall": clean_val(row.get("태양광 설치 지점")) in ["○", "O", "o", "1", "Y", "y"],
            "refWaterLevel": clean_val(row.get("기준수위")),
            "waterLevelType": clean_val(row.get("수위계 방식")),
            "memo": clean_val(row.get("비고")),
            "coords": {
                "lonDMS": lon_raw,
                "latDMS": lat_raw,
                "lon": lon_dd,
                "lat": lat_dd
            },
            "maintenance": {
                "mountType": m_info.get("설치방식", ""),
                "stationType": m_info.get("국사형태", ""),
                "waterLevelInstalled": m_info.get("수위계 설치 여부", ""),
                "batteryStatus": m_info.get("노후 배터리", ""),
                "signStatus": m_info.get("점용표지판(낙동강)", "") or m_info.get("점용표지판(영산강)", ""),
                "boardStatus": m_info.get("관측소 현황판", ""),
                "cctv": m_info.get("CCTV", ""),
                "nvr": m_info.get("NVR", ""),
                "solarDetail": m_info.get("태양광", ""),
                "history": []
            }
        }
        stations.append(station_obj)

    data_dir = os.path.join(system_dir, "data")
    os.makedirs(data_dir, exist_ok=True)
    
    json_path = os.path.join(data_dir, "stations_initial.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(stations, f, ensure_ascii=False, indent=2)
    print(f"Saved {len(stations)} stations to {json_path}")

    js_path = os.path.join(data_dir, "stations_initial.js")
    with open(js_path, "w", encoding="utf-8") as f:
        f.write("window.INITIAL_STATIONS_DATA = " + json.dumps(stations, ensure_ascii=False, indent=2) + ";\n")
    print(f"Saved JS bundle to {js_path}")

if __name__ == "__main__":
    main()
